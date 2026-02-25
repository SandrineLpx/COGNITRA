from __future__ import annotations

import hashlib
import importlib.util
import json
import re
import zipfile
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from xml.sax.saxutils import escape as xml_escape

import pandas as pd

from src.constants import (
    CANON_TOPICS,
    DISPLAY_REGIONS,
    FOOTPRINT_REGIONS,
    FOOTPRINT_TO_DISPLAY,
    MACRO_THEME_RULES,
    PREMIUM_OEMS,
    UNCERTAINTY_WORDS,
)
from src.pdf_extract import extract_text_robust
from src.dedupe import build_dedupe_key, _title_fingerprint
from src.postprocess import (
    COUNTRY_TO_FOOTPRINT,
    _COMPANY_SPECIAL_CANONICAL,
    _KEY_OEMS,
    _LEGAL_SUFFIX_TOKENS,
    _OEM_CANONICAL_BY_LOWER,
)
from src.storage import RECORDS_PATH, load_records

DATA_DIR = RECORDS_PATH.parent
BRIEFS_DIR = DATA_DIR / "briefs"
BRIEF_INDEX = BRIEFS_DIR / "index.jsonl"

QUALITY_DIR = DATA_DIR / "quality"
RECORD_QC_LOG = QUALITY_DIR / "record_qc.jsonl"
BRIEF_QC_LOG = QUALITY_DIR / "brief_qc.jsonl"
QUALITY_RUNS_LOG = QUALITY_DIR / "quality_runs.jsonl"
QUALITY_REPORT_XLSX = QUALITY_DIR / "quality_report.xlsx"

_UNCERTAINTY_RE = re.compile(UNCERTAINTY_WORDS, re.IGNORECASE)
_SOFT_LANGUAGE_RE = _UNCERTAINTY_RE  # same word list; alias kept for readability
_OVERREACH_RE = re.compile(
    r"\b(decided|final decision|scrapping|scrap(?:ped)?|will scrap|has scrapped)\b",
    re.IGNORECASE,
)
_REC_REF_RE = re.compile(r"\bREC\s*[:#]\s*([A-Za-z0-9_-]+)\b", re.IGNORECASE)
_TITLE_HEAD_RE = re.compile(r"^\s*#*\s*([A-Z][A-Z0-9 &/\-()]+)\s*$")
_SUMMARY_HEAD_RE = re.compile(r"<summary>\s*([^<]+?)\s*</summary>", re.IGNORECASE)
_WEEK_RANGE_RE = re.compile(r"\blast\s+(\d+)\s+days\b", re.IGNORECASE)

_CLAIM_HEADINGS = {
    "EXECUTIVE SUMMARY",
    "HIGH PRIORITY DEVELOPMENTS",
    "FOOTPRINT REGION SIGNALS",
    "KEY DEVELOPMENTS BY TOPIC",
    "EMERGING TRENDS",
    "CONFLICTS & UNCERTAINTY",
}

_KNOWN_HEADINGS = {
    "EXECUTIVE SUMMARY",
    "HIGH PRIORITY DEVELOPMENTS",
    "FOOTPRINT REGION SIGNALS",
    "KEY DEVELOPMENTS BY TOPIC",
    "EMERGING TRENDS",
    "CONFLICTS & UNCERTAINTY",
    "RECOMMENDED ACTIONS",
    "APPENDIX",
}

_STOPWORDS = {
    "the",
    "a",
    "an",
    "and",
    "or",
    "but",
    "in",
    "on",
    "at",
    "to",
    "for",
    "of",
    "by",
    "with",
    "from",
    "as",
    "is",
    "are",
    "was",
    "were",
    "be",
    "been",
    "that",
    "this",
    "it",
    "its",
    "into",
    "than",
    "then",
    "their",
    "they",
}
_CANON_TOPICS_LOWER = {str(t).strip().lower() for t in CANON_TOPICS if str(t).strip()}

def _build_company_alias_groups() -> Dict[str, set]:
    """Auto-derive company alias groups from postprocess canonicalization maps."""
    groups: Dict[str, set] = {}
    # Invert _OEM_CANONICAL_BY_LOWER: group all keys mapping to same canonical
    for alias, canonical in _OEM_CANONICAL_BY_LOWER.items():
        key = canonical.lower()
        groups.setdefault(key, set()).add(alias)
        groups[key].add(key)
    # Merge _COMPANY_SPECIAL_CANONICAL (strip legal suffixes → same canonical)
    for alias, canonical in _COMPANY_SPECIAL_CANONICAL.items():
        key = canonical.lower()
        groups.setdefault(key, set()).add(alias)
        groups[key].add(key)
    # Add _KEY_OEMS entries that resolve to existing groups or form new ones
    for oem in _KEY_OEMS:
        mapped = _OEM_CANONICAL_BY_LOWER.get(oem)
        if mapped:
            groups.setdefault(mapped.lower(), set()).add(oem)
        else:
            groups.setdefault(oem, set()).add(oem)
    # Add PREMIUM_OEMS entries
    for oem in PREMIUM_OEMS:
        mapped = _OEM_CANONICAL_BY_LOWER.get(oem)
        if mapped:
            groups.setdefault(mapped.lower(), set()).add(oem)
        else:
            groups.setdefault(oem, set()).add(oem)
    # Common OEM aliases not yet in postprocess maps but frequently seen in articles
    _EXTRA_ALIASES: Dict[str, set] = {
        "general motors": {"gm", "general motors", "general motors co"},
        "stellantis": {"stellantis", "stellantis nv", "stellantis n.v."},
        "hyundai": {"hyundai", "hyundai motor", "hyundai motor group", "hyundai motor company"},
        "ford": {"ford", "ford motor", "ford motor company", "ford motor co"},
        "nissan": {"nissan", "nissan motor", "nissan motor co"},
        "honda": {"honda", "honda motor", "honda motor co"},
        "renault": {"renault", "renault group", "renault sa"},
        "geely": {"geely", "geely auto", "geely automobile"},
        "tata": {"tata", "tata motors", "tata motors ltd"},
        "volvo cars": {"volvo", "volvo cars", "volvo car group"},
        "jaguar": {"jaguar", "jaguar land rover", "jlr"},
        "kia": {"kia", "kia motors", "kia corporation"},
        "audi": {"audi", "audi ag"},
        "porsche": {"porsche", "porsche ag", "porsche automobil"},
    }
    for canonical, variants in _EXTRA_ALIASES.items():
        existing = groups.get(canonical, set())
        groups[canonical] = existing | variants
    # Only keep groups with 2+ variants (single-entry groups can't have inconsistencies)
    return {k: v for k, v in groups.items() if len(v) >= 2}


_COMPANY_ALIAS_GROUPS = _build_company_alias_groups()


def _now() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def _now_iso() -> str:
    return _now().isoformat()


def _new_run_id() -> str:
    return f"qc_{_now().strftime('%Y%m%d_%H%M%S')}"


def _ensure_quality_dir() -> None:
    QUALITY_DIR.mkdir(parents=True, exist_ok=True)


def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    out: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if isinstance(obj, dict):
                out.append(obj)
    return out


def _next_run_version() -> int:
    rows = _read_jsonl(QUALITY_RUNS_LOG)
    versions: List[int] = []
    for row in rows:
        try:
            value = int(row.get("run_version"))
        except Exception:
            continue
        if value > 0:
            versions.append(value)
    if versions:
        return max(versions) + 1
    return len(rows) + 1


def _append_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> int:
    rows_list = [r for r in rows if isinstance(r, dict)]
    if not rows_list:
        return 0
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        for row in rows_list:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    return len(rows_list)


def _parse_date(value: Any) -> Optional[datetime]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d",):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def _record_date(rec: Dict[str, Any]) -> Optional[datetime]:
    d = _parse_date(rec.get("publish_date"))
    if d:
        return d
    return _parse_date(rec.get("created_at"))


def _parse_week_range_days(week_range: Optional[str], default_days: int = 7) -> int:
    if not isinstance(week_range, str):
        return default_days
    m = _WEEK_RANGE_RE.search(week_range)
    if not m:
        return default_days
    try:
        days = int(m.group(1))
    except Exception:
        return default_days
    return max(1, min(365, days))


def _norm_text(value: Any) -> str:
    s = str(value or "").lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _to_str_list(value: Any) -> List[str]:
    if not isinstance(value, list):
        return []
    out: List[str] = []
    for v in value:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.append(s)
    return out


def _dup_values(values: Sequence[str]) -> List[str]:
    seen: Dict[str, str] = {}
    dups: List[str] = []
    for val in values:
        k = _norm_text(val)
        if not k:
            continue
        if k in seen and seen[k] not in dups:
            dups.append(seen[k])
        else:
            seen[k] = str(val)
    return dups


def _keywords(text: str) -> set[str]:
    toks = _norm_text(text).split()
    return {t for t in toks if len(t) >= 4 and t not in _STOPWORDS}


def _is_bullet_line(line: str) -> bool:
    s = str(line).lstrip()
    return s.startswith("-") or s.startswith("*") or s.startswith("•")


def _strip_topic_label_candidate(line: str) -> str:
    s = re.sub(r"^\s*[-*•]\s*", "", str(line or "")).strip()
    s = s.replace("*", "").replace("_", "").replace("`", "").strip()
    if s.endswith(":"):
        s = s[:-1].strip()
    return s


def _is_structural_topic_label_bullet(line: str, section: str) -> bool:
    if section != "KEY DEVELOPMENTS BY TOPIC" or not _is_bullet_line(line):
        return False
    candidate = _strip_topic_label_candidate(line).lower()
    return bool(candidate) and candidate in _CANON_TOPICS_LOWER


def _rec_refs(text: str) -> List[str]:
    return [m.group(1) for m in _REC_REF_RE.finditer(text or "")]


def _resolve_rec_refs(
    refs: Sequence[str],
    selected_record_ids: Sequence[str],
) -> Tuple[List[str], List[Tuple[str, str]]]:
    """Resolve REC refs against selected_record_ids.

    Supports legacy numeric REC labels (REC:1..N) by mapping them to
    selected_record_ids in list order.
    """
    ordered_ids = [str(x) for x in selected_record_ids if str(x)]
    valid_set = set(ordered_ids)
    resolved: List[str] = []
    mapped_pairs: List[Tuple[str, str]] = []

    for raw in refs:
        ref = str(raw or "").strip()
        if not ref:
            continue
        if ref in valid_set:
            resolved.append(ref)
            continue
        if ref.isdigit() and ordered_ids:
            idx = int(ref) - 1
            if 0 <= idx < len(ordered_ids):
                mapped = ordered_ids[idx]
                resolved.append(mapped)
                mapped_pairs.append((ref, mapped))
                continue
        resolved.append(ref)
    return resolved, mapped_pairs


def _hash_claim(text: str) -> str:
    return hashlib.sha1((text or "").encode("utf-8")).hexdigest()[:12]


def _extract_sentence(text: str, at: int) -> str:
    if not text:
        return ""
    left = text.rfind("\n", 0, at)
    left = 0 if left < 0 else left + 1
    right = text.find("\n", at)
    right = len(text) if right < 0 else right
    return text[left:right].strip()


def _latest_brief_file() -> Optional[Path]:
    if not BRIEFS_DIR.exists():
        return None
    files = sorted(
        [p for p in BRIEFS_DIR.glob("brief_*.md") if p.is_file()],
        key=lambda p: (p.stat().st_mtime, p.name),
    )
    return files[-1] if files else None


def _brief_file_from_id(brief_id: Optional[str]) -> Optional[Path]:
    if not brief_id:
        return None
    candidate = Path(str(brief_id))
    if candidate.exists() and candidate.is_file():
        return candidate
    bid = str(brief_id).strip()
    if not bid:
        return None
    if not bid.endswith(".md"):
        bid = f"{bid}.md"
    if bid.startswith("brief_"):
        p = BRIEFS_DIR / bid
        if p.exists() and p.is_file():
            return p
    p2 = BRIEFS_DIR / bid
    if p2.exists() and p2.is_file():
        return p2
    return None


def _brief_meta_for_file(brief_path: Optional[Path]) -> Dict[str, Any]:
    if not brief_path:
        return {}
    sidecar = brief_path.with_suffix(".meta.json")
    if sidecar.exists():
        try:
            obj = json.loads(sidecar.read_text(encoding="utf-8"))
            if isinstance(obj, dict):
                return obj
        except Exception:
            pass

    rows = _read_jsonl(BRIEF_INDEX)
    if not rows:
        return {}
    name = brief_path.name
    matches = []
    for row in rows:
        file_field = str(row.get("file") or "")
        file_name = Path(file_field.replace("\\", "/")).name
        if file_name == name:
            matches.append(row)
    return matches[-1] if matches else rows[-1]


def _select_target_records(
    records: List[Dict[str, Any]],
    selected_ids: Sequence[str],
    week_days: int,
) -> Tuple[List[Dict[str, Any]], str]:
    by_id = {str(r.get("record_id") or ""): r for r in records}
    if selected_ids:
        picked = [by_id[rid] for rid in selected_ids if rid in by_id]
        return picked, "brief_selected_ids"

    cutoff = _now() - timedelta(days=week_days)
    approved: List[Dict[str, Any]] = []
    for rec in records:
        if str(rec.get("review_status") or "") != "Approved":
            continue
        rd = _record_date(rec)
        if rd and rd >= cutoff:
            approved.append(rec)
    return approved, "approved_in_week_range"


def _record_finding(
    run_id: str,
    rec: Dict[str, Any],
    finding_type: str,
    field: str,
    severity: str,
    grounded: str,
    impact: str,
    notes: str,
) -> Dict[str, Any]:
    return {
        "run_id": run_id,
        "created_at": _now_iso(),
        "record_id": str(rec.get("record_id") or ""),
        "doc_title": str(rec.get("title") or "Untitled"),
        "version": None,
        "finding_type": finding_type,
        "field": field,
        "severity": severity,
        "grounded": grounded,
        "impact": impact,
        "notes": notes,
        "status": "open",
    }


def _brief_finding(
    run_id: str,
    brief_id: str,
    section: str,
    claim_text: str,
    supported_record_ids: Sequence[str],
    grounded_to_records: bool,
    issue_type: str,
    severity: str,
    notes: str,
) -> Dict[str, Any]:
    return {
        "run_id": run_id,
        "created_at": _now_iso(),
        "version": None,
        "brief_id": brief_id,
        "section": section,
        "claim_id": _hash_claim(claim_text),
        "claim_text": claim_text,
        "supported_record_ids": list(supported_record_ids),
        "grounded_to_records": bool(grounded_to_records),
        "issue_type": issue_type,
        "severity": severity,
        "notes": notes,
        "status": "open",
    }


def _extract_brief_sections(text: str) -> Tuple[Dict[str, str], List[Tuple[str, str]]]:
    sections: Dict[str, List[str]] = {}
    lined: List[Tuple[str, str]] = []
    current = "UNKNOWN"
    sections[current] = []
    for raw in (text or "").splitlines():
        line = raw.rstrip("\n")
        stripped = line.strip()
        m = _TITLE_HEAD_RE.match(stripped)
        if m:
            cand = m.group(1).strip().upper()
            if cand in _KNOWN_HEADINGS:
                current = cand
                sections.setdefault(current, [])
                continue
        # Also support saved brief format that wraps sections in HTML details/summary tags.
        sm = _SUMMARY_HEAD_RE.search(stripped)
        if sm:
            cand = re.sub(r"\s+", " ", sm.group(1)).strip().upper()
            if cand in _KNOWN_HEADINGS:
                current = cand
                sections.setdefault(current, [])
                tail = stripped[sm.end():].strip()
                if tail:
                    sections.setdefault(current, []).append(tail)
                    lined.append((current, tail))
                continue
        if re.fullmatch(r"</?details>\s*", stripped, re.IGNORECASE):
            continue
        sections.setdefault(current, []).append(line)
        lined.append((current, line))
    return {k: "\n".join(v).strip() for k, v in sections.items()}, lined


def _record_contains_uncertainty(rec: Dict[str, Any]) -> bool:
    if str(rec.get("confidence") or "") in {"Medium", "Low"}:
        return True
    text = " ".join(_to_str_list(rec.get("evidence_bullets")) + _to_str_list(rec.get("key_insights")))
    return bool(_UNCERTAINTY_RE.search(text))


def _records_have_soft_language(records: Sequence[Dict[str, Any]]) -> bool:
    for rec in records:
        text = " ".join(_to_str_list(rec.get("evidence_bullets")) + _to_str_list(rec.get("key_insights")))
        if _SOFT_LANGUAGE_RE.search(text):
            return True
    return False


def _cross_record_theme_count(lined: Sequence[Tuple[str, str]], valid_ids: set[str]) -> int:
    count = 0
    for section, line in lined:
        if section not in {"EXECUTIVE SUMMARY", "KEY DEVELOPMENTS BY TOPIC", "EMERGING TRENDS"}:
            continue
        if not _is_bullet_line(line):
            continue
        refs = set(_rec_refs(line))
        if valid_ids:
            refs = {r for r in refs if r in valid_ids}
        if len(refs) >= 2:
            count += 1
    return count


def _action_specificity_score(actions_text: str) -> int:
    t = str(actions_text or "")
    checks = [
        bool(re.search(r"\b(owner|vp|director|procurement|engineering|strategy|sales)\b", t, re.IGNORECASE)),
        bool(re.search(r"\b(immediate|quarter|month|months|week|weeks|6 months|next)\b", t, re.IGNORECASE)),
        bool(re.search(r"\b(develop|conduct|prepare|update|create|build|assess|review|draft|deliver)\b", t, re.IGNORECASE)),
        bool(re.search(r"\b(trigger|watch|if|when|until|revisit)\b", t, re.IGNORECASE)),
        bool(re.search(r"\b(forecast|memo|playbook|report|list|dashboard|risk|plan)\b", t, re.IGNORECASE)),
    ]
    score = sum(1 for ok in checks if ok)
    return max(1, min(5, score))


EVIDENCE_GROUNDING_THRESHOLD = 0.60
# Set equal to the grounding threshold to enforce strict 60% cutoff:
# overlap < 60% => hard miss.
EVIDENCE_NEAR_MISS_THRESHOLD = 0.60

_DISPLAY_SET = set(DISPLAY_REGIONS)
_FOOTPRINT_SET = set(FOOTPRINT_REGIONS)

# Parent-bucket coverage: if a country is "missing" but its broader bucket
# is present, the region is functionally covered (e.g. "West Europe" covers
# "Germany").  Used to suppress false-positive missing_footprint_region.
_COUNTRY_PARENT_BUCKETS: Dict[str, set[str]] = {
    "Germany": {"West Europe", "Europe"},
    "France": {"West Europe", "Europe"},
    "Italy": {"West Europe", "Europe"},
    "Spain": {"West Europe", "Europe"},
    "Portugal": {"West Europe", "Europe"},
    "Sweden": {"West Europe", "Europe"},
    "United Kingdom": {"West Europe", "Europe"},
    "Czech Republic": {"Central Europe", "Europe"},
    "Russia": {"East Europe", "Europe"},
    "Mexico": {"NAFTA"},
    "United States": {"NAFTA"},
    "Thailand": {"ASEAN", "South Asia"},
    "India": {"Indian Subcontinent", "South Asia"},
    "China": {"South Asia"},
    "Taiwan": {"South Asia"},
    "Japan": {"South Asia"},
    "South Korea": {"South Asia"},
    "Morocco": {"Africa"},
}


def _check_macro_themes(
    run_id: str,
    rec: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Validate macro theme detection audit fields against rule definitions."""
    findings: List[Dict[str, Any]] = []
    detail = rec.get("_macro_theme_detail") or {}
    detected = rec.get("macro_themes_detected") or []
    companies_l = {c.lower() for c in (rec.get("companies_mentioned") or []) if isinstance(c, str)}
    regions_set = set(rec.get("regions_mentioned") or []) | set(rec.get("regions_relevant_to_apex_mobility") or [])

    for rule in MACRO_THEME_RULES:
        name = rule["name"]
        theme_detail = detail.get(name, {})
        fired = theme_detail.get("fired", name in detected)

        if not fired:
            continue

        min_groups = rule["min_groups"]
        groups_matched = theme_detail.get("groups_matched") or []

        # Check min_groups actually met
        if len(groups_matched) < min_groups:
            findings.append(_record_finding(
                run_id, rec,
                finding_type="macro_theme_min_groups_violation",
                field="macro_themes_detected",
                severity="High",
                grounded="Yes",
                impact="Theme fired without meeting minimum signal group threshold.",
                notes=f"'{name}' requires {min_groups} groups, only {len(groups_matched)} matched: {groups_matched}",
            ))

        # Check premium company gate
        if rule.get("premium_company_gate") and not (companies_l & PREMIUM_OEMS):
            findings.append(_record_finding(
                run_id, rec,
                finding_type="macro_theme_premium_gate_violation",
                field="macro_themes_detected",
                severity="High",
                grounded="Yes",
                impact="Theme requires a premium OEM but none found in companies_mentioned.",
                notes=f"'{name}' has premium_company_gate=True but companies={list(companies_l)}",
            ))

        # Check region requirements
        req_regions = rule.get("region_requirements")
        if req_regions and not (regions_set & req_regions):
            findings.append(_record_finding(
                run_id, rec,
                finding_type="macro_theme_region_requirement_violation",
                field="macro_themes_detected",
                severity="High",
                grounded="Yes",
                impact="Theme requires specific regions but none present.",
                notes=f"'{name}' requires one of {sorted(req_regions)}, record has {sorted(regions_set)}",
            ))

    return findings


def _check_geo_completeness(
    run_id: str,
    rec: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], bool]:
    """Extended geo validation: missing regions, display bucket leakage."""
    findings: List[Dict[str, Any]] = []
    geo_ok = True

    countries = _to_str_list(rec.get("country_mentions"))
    footprint_regions = _to_str_list(rec.get("regions_relevant_to_apex_mobility"))
    display_regions = _to_str_list(rec.get("regions_mentioned"))

    # In the new region design FOOTPRINT_REGIONS == DISPLAY_REGIONS (same set),
    # so every valid footprint value is also a valid display value.
    # Check that all values in regions_mentioned are valid FOOTPRINT_REGIONS entries.
    invalid_display = [r for r in display_regions if r not in _FOOTPRINT_SET]
    if invalid_display:
        geo_ok = False
        findings.append(_record_finding(
            run_id, rec,
            finding_type="invalid_display_region",
            field="regions_mentioned",
            severity="High",
            grounded="Yes",
            impact="Unknown region value in regions_mentioned — not in FOOTPRINT_REGIONS.",
            notes=f"Invalid values in regions_mentioned: {invalid_display}. Must be one of the canonical FOOTPRINT_REGIONS entries.",
        ))

    # Check for missing footprint regions when countries are present.
    # Exclude regions that postprocess deliberately removed (tracked in
    # _region_validation_flags, e.g. "us_region_removed_no_us_evidence").
    expected_from_countries = list(dict.fromkeys(
        COUNTRY_TO_FOOTPRINT[c] for c in countries if c in COUNTRY_TO_FOOTPRINT
    ))
    if expected_from_countries:
        footprint_set = set(footprint_regions)
        validation_flags = rec.get("_region_validation_flags") or []
        deliberately_removed: set[str] = set()
        for flag in validation_flags:
            flag_s = str(flag)
            if "us_region_removed" in flag_s:
                deliberately_removed.add("United States")
            if "china_region_removed" in flag_s:
                deliberately_removed.add("China")
        missing = [
            r for r in expected_from_countries
            if r not in footprint_set
            and r not in deliberately_removed
            and not (footprint_set & _COUNTRY_PARENT_BUCKETS.get(r, set()))
        ]
        if missing:
            geo_ok = False
            findings.append(_record_finding(
                run_id, rec,
                finding_type="missing_footprint_region",
                field="regions_relevant_to_apex_mobility",
                severity="Medium",
                grounded="Yes",
                impact="Country present but derived footprint region missing.",
                notes=f"Countries {countries} imply footprint regions {expected_from_countries}, missing: {missing}",
            ))

    return findings, geo_ok


def _check_confidence_alignment(
    run_id: str,
    rec: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Validate confidence aligns with evidence density and source quality."""
    findings: List[Dict[str, Any]] = []
    confidence = str(rec.get("confidence") or "")
    evidence = _to_str_list(rec.get("evidence_bullets"))
    source_type = str(rec.get("source_type") or "Other")
    evidence_count = len(evidence)

    # High confidence with thin evidence
    if confidence == "High" and evidence_count < 2:
        findings.append(_record_finding(
            run_id, rec,
            finding_type="confidence_evidence_mismatch",
            field="confidence",
            severity="Medium",
            grounded="Yes",
            impact="High confidence with thin evidence may overstate reliability.",
            notes=f"confidence=High but only {evidence_count} evidence bullet(s).",
        ))

    # High confidence from unknown source
    if confidence == "High" and source_type == "Other":
        findings.append(_record_finding(
            run_id, rec,
            finding_type="confidence_source_mismatch",
            field="confidence",
            severity="Low",
            grounded="Yes",
            impact="High confidence from unrecognized source type may be inflated.",
            notes=f"confidence=High but source_type='Other'.",
        ))

    return findings


def _check_priority_reason(
    run_id: str,
    rec: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Validate priority_reason exists and is non-empty when priority was escalated."""
    findings: List[Dict[str, Any]] = []
    priority = str(rec.get("priority") or "")
    priority_reason = str(rec.get("priority_reason") or "").strip()
    priority_llm = str(rec.get("priority_llm") or "").strip()

    # If priority was escalated (differs from LLM original), reason should exist
    if priority_llm and priority != priority_llm and not priority_reason:
        findings.append(_record_finding(
            run_id, rec,
            finding_type="missing_priority_reason",
            field="priority_reason",
            severity="Medium",
            grounded="Yes",
            impact="Priority escalation without explanation reduces audit transparency.",
            notes=f"priority escalated from '{priority_llm}' to '{priority}' but priority_reason is empty.",
        ))

    # High priority should always have a reason
    if priority == "High" and not priority_reason:
        findings.append(_record_finding(
            run_id, rec,
            finding_type="missing_priority_reason",
            field="priority_reason",
            severity="Low",
            grounded="Yes",
            impact="High priority without explanation reduces audit transparency.",
            notes="priority=High but priority_reason is empty.",
        ))

    return findings


def _check_duplicate_records(
    run_id: str,
    records: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Post-hoc duplicate detection across target records using dedupe keys and fuzzy titles."""
    from difflib import SequenceMatcher

    findings: List[Dict[str, Any]] = []
    if len(records) < 2:
        return findings

    # Group by dedupe key
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for rec in records:
        key = build_dedupe_key(rec)
        groups.setdefault(key, []).append(rec)
    for key, group in groups.items():
        if len(group) < 2:
            continue
        ids = [str(r.get("record_id") or "") for r in group]
        for rec in group[1:]:
            findings.append(_record_finding(
                run_id, rec,
                finding_type="duplicate_record_in_brief",
                field="title",
                severity="High",
                grounded="Yes",
                impact="Duplicate records inflate topic counts and may double-count in brief.",
                notes=f"Same dedupe_key as records: {ids}. Key: {key[:80]}",
            ))

    # Fuzzy title matching across all pairs (O(n^2) but n is typically <50)
    indexed: List[Tuple[Dict[str, Any], str]] = []
    for rec in records:
        fp = _title_fingerprint(rec.get("title") or "")
        if fp:
            indexed.append((rec, fp))
    seen_pairs: set = set()
    for i, (rec_a, fp_a) in enumerate(indexed):
        for j in range(i + 1, len(indexed)):
            rec_b, fp_b = indexed[j]
            pair_key = tuple(sorted([
                str(rec_a.get("record_id") or ""),
                str(rec_b.get("record_id") or ""),
            ]))
            if pair_key in seen_pairs:
                continue
            ratio = SequenceMatcher(None, fp_a, fp_b).ratio()
            if ratio >= 0.85:
                seen_pairs.add(pair_key)
                # Skip if already caught by exact dedupe key
                key_a = build_dedupe_key(rec_a)
                key_b = build_dedupe_key(rec_b)
                if key_a == key_b:
                    continue
                findings.append(_record_finding(
                    run_id, rec_b,
                    finding_type="near_duplicate_titles",
                    field="title",
                    severity="Medium",
                    grounded="Yes",
                    impact="Near-duplicate titles may represent the same story from different sources.",
                    notes=f"similarity={ratio:.0%} with '{rec_a.get('title', '')[:60]}' (id={rec_a.get('record_id')})",
                ))
    return findings


def run_record_qc(
    run_id: str,
    records: Sequence[Dict[str, Any]],
    run_version: int,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    findings: List[Dict[str, Any]] = []
    evidence_pass: Dict[str, bool] = {}
    canonical_pass: Dict[str, bool] = {}
    geo_pass: Dict[str, bool] = {}
    pdf_cache: Dict[str, Tuple[str, str]] = {}

    def load_pdf_text(path_str: str) -> Tuple[str, str]:
        if path_str in pdf_cache:
            return pdf_cache[path_str]
        path = Path(path_str)
        if not path.exists() or not path.is_file():
            pdf_cache[path_str] = ("", "missing")
            return pdf_cache[path_str]
        try:
            text, method = extract_text_robust(path.read_bytes())
            pdf_cache[path_str] = (text or "", method or "unknown")
        except Exception:
            pdf_cache[path_str] = ("", "error")
        return pdf_cache[path_str]

    for rec in records:
        rid = str(rec.get("record_id") or "")
        if not rid:
            continue
        canonical_ok = True
        geo_ok = True
        all_evidence_grounded = True

        for field in ("companies_mentioned", "country_mentions", "regions_relevant_to_apex_mobility"):
            vals = _to_str_list(rec.get(field))
            dups = _dup_values(vals)
            if dups:
                findings.append(
                    _record_finding(
                        run_id,
                        rec,
                        finding_type="duplicate_values",
                        field=field,
                        severity="Medium",
                        grounded="Yes",
                        impact="Duplicate values can bias themes and summary counts.",
                        notes=f"Duplicate values detected: {', '.join(dups)}",
                    )
                )

        countries = _to_str_list(rec.get("country_mentions"))
        regions = _to_str_list(rec.get("regions_relevant_to_apex_mobility"))
        expected_regions = list(dict.fromkeys(COUNTRY_TO_FOOTPRINT[c] for c in countries if c in COUNTRY_TO_FOOTPRINT))

        # Only flag regions that are NOT valid footprint values at all.
        # Postprocess intentionally adds text-hinted parent-bucket regions
        # (e.g. "West Europe" when only "Germany" is in country_mentions),
        # so extra-but-valid footprint regions are not leakage.
        invalid_regions = [r for r in regions if r not in _FOOTPRINT_SET]
        if invalid_regions:
            geo_ok = False
            findings.append(
                _record_finding(
                    run_id,
                    rec,
                    finding_type="geo_leakage",
                    field="regions_relevant_to_apex_mobility",
                    severity="High",
                    grounded="Yes",
                    impact="Footprint signals may be distorted in brief rollups.",
                    notes=f"Invalid footprint values (not in FOOTPRINT_REGIONS): {invalid_regions}",
                )
            )

        companies = _to_str_list(rec.get("companies_mentioned"))
        company_norm = [_norm_text(c) for c in companies]
        alias_inconsistencies: List[str] = []
        alias_noncanonical: List[str] = []
        for canonical, variants in _COMPANY_ALIAS_GROUPS.items():
            present = [c for c, n in zip(companies, company_norm) if n in variants]
            present_norm = {n for n in company_norm if n in variants}
            if len(present_norm) > 1:
                alias_inconsistencies.append(f"{canonical}: {present}")
            elif len(present_norm) == 1 and canonical not in present_norm:
                alias_noncanonical.append(f"{canonical}: {present[0]}")
        if alias_inconsistencies:
            canonical_ok = False
            findings.append(
                _record_finding(
                    run_id,
                    rec,
                    finding_type="canonicalization_inconsistent",
                    field="companies_mentioned",
                    severity="Medium",
                    grounded="Yes",
                    impact="Entity rollups become unstable across records.",
                    notes=" | ".join(alias_inconsistencies),
                )
            )
        elif alias_noncanonical:
            findings.append(
                _record_finding(
                    run_id,
                    rec,
                    finding_type="canonicalization_alias_present",
                    field="companies_mentioned",
                    severity="Low",
                    grounded="Yes",
                    impact="Alias form can reduce grouping consistency.",
                    notes=" | ".join(alias_noncanonical),
                )
            )

        evidence = _to_str_list(rec.get("evidence_bullets"))
        if not evidence:
            all_evidence_grounded = False
            findings.append(
                _record_finding(
                    run_id,
                    rec,
                    finding_type="missing_evidence",
                    field="evidence_bullets",
                    severity="High",
                    grounded="No",
                    impact="Record cannot be validated for factual support.",
                    notes="evidence_bullets is empty.",
                )
            )
        else:
            pdf_path = str(rec.get("source_pdf_path") or "").strip()
            if not pdf_path:
                all_evidence_grounded = False
                findings.append(
                    _record_finding(
                        run_id,
                        rec,
                        finding_type="evidence_check_skipped",
                        field="evidence_bullets",
                        severity="Medium",
                        grounded="Partial",
                        impact="Grounding coverage reduced; PDF unavailable.",
                        notes="source_pdf_path missing; skipped text grounding check.",
                    )
                )
            else:
                pdf_text, method = load_pdf_text(pdf_path)
                if len(pdf_text.strip()) < 100:
                    all_evidence_grounded = False
                    findings.append(
                        _record_finding(
                            run_id,
                            rec,
                            finding_type="evidence_check_skipped",
                            field="evidence_bullets",
                            severity="Medium",
                            grounded="Partial",
                            impact="Grounding coverage reduced; extracted PDF text is unavailable/short.",
                            notes=f"PDF read method={method}; extracted text too short for reliable check.",
                        )
                    )
                else:
                    norm_doc = _norm_text(pdf_text)
                    doc_terms = _keywords(pdf_text)
                    hard_misses: List[Tuple[str, float]] = []
                    near_misses: List[Tuple[str, float]] = []
                    overlap_scores: List[float] = []
                    for bullet in evidence:
                        norm_bullet = _norm_text(bullet)
                        if norm_bullet and norm_bullet in norm_doc:
                            overlap_scores.append(1.0)
                            continue
                        terms = _keywords(bullet)
                        if not terms:
                            overlap_scores.append(1.0)
                            continue
                        overlap = len(terms & doc_terms) / max(1, len(terms))
                        overlap_scores.append(overlap)
                        if overlap >= EVIDENCE_GROUNDING_THRESHOLD:
                            continue
                        if overlap >= EVIDENCE_NEAR_MISS_THRESHOLD:
                            near_misses.append((bullet, overlap))
                        else:
                            hard_misses.append((bullet, overlap))
                    avg_overlap = sum(overlap_scores) / max(1, len(overlap_scores))
                    if hard_misses:
                        all_evidence_grounded = False
                        preview = " | ".join(
                            f"{b[:80]}(overlap={o:.0%})" for b, o in hard_misses[:2]
                        )
                        findings.append(
                            _record_finding(
                                run_id,
                                rec,
                                finding_type="evidence_not_grounded",
                                field="evidence_bullets",
                                severity="High",
                                grounded="No",
                                impact="Ungrounded evidence can propagate factual errors to briefs.",
                                notes=f"{len(hard_misses)} bullet(s) below {EVIDENCE_GROUNDING_THRESHOLD:.0%} threshold. avg_overlap={avg_overlap:.0%}. Samples: {preview}",
                            )
                        )
                    if near_misses:
                        preview = " | ".join(
                            f"{b[:80]}(overlap={o:.0%})" for b, o in near_misses[:2]
                        )
                        findings.append(
                            _record_finding(
                                run_id,
                                rec,
                                finding_type="evidence_near_miss",
                                field="evidence_bullets",
                                severity="Medium",
                                grounded="Partial",
                                impact="Evidence may be paraphrased; verify factual accuracy.",
                                notes=f"{len(near_misses)} bullet(s) between {EVIDENCE_NEAR_MISS_THRESHOLD:.0%}-{EVIDENCE_GROUNDING_THRESHOLD:.0%}. avg_overlap={avg_overlap:.0%}. Samples: {preview}",
                            )
                        )

        # --- Extended geo validation (issue #3) ---
        geo_ext_findings, geo_ext_ok = _check_geo_completeness(run_id, rec)
        findings.extend(geo_ext_findings)
        if not geo_ext_ok:
            geo_ok = False

        # --- Macro theme verification (issue #2) ---
        findings.extend(_check_macro_themes(run_id, rec))

        # --- Confidence alignment (issue #4) ---
        findings.extend(_check_confidence_alignment(run_id, rec))

        # --- Priority reason audit (issue #5) ---
        findings.extend(_check_priority_reason(run_id, rec))

        evidence_pass[rid] = all_evidence_grounded
        canonical_pass[rid] = canonical_ok
        geo_pass[rid] = geo_ok

    # --- Cross-record duplicate detection ---
    findings.extend(_check_duplicate_records(run_id, list(records)))

    for finding in findings:
        finding["version"] = run_version

    all_rids = [str(r.get("record_id") or "") for r in records if r.get("record_id")]
    per_record = _per_record_scores(findings, all_rids)

    metrics = {
        "evidence_pass": evidence_pass,
        "canonical_pass": canonical_pass,
        "geo_pass": geo_pass,
        "per_record_scores": per_record,
    }
    return findings, metrics


def run_brief_qc(
    run_id: str,
    run_version: int,
    brief_id: str,
    brief_text: str,
    selected_record_ids: Sequence[str],
    selected_records: Sequence[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    findings: List[Dict[str, Any]] = []
    sections, lined = _extract_brief_sections(brief_text or "")
    valid_ids = {str(x) for x in selected_record_ids if str(x)}
    has_rec_labels = bool(_REC_REF_RE.search(brief_text or ""))

    if has_rec_labels:
        for section, line in lined:
            text = line.strip()
            if not text:
                continue
            refs_raw = _rec_refs(text)
            refs, mapped_pairs = _resolve_rec_refs(refs_raw, selected_record_ids)
            if refs:
                invalid = [r for r in refs if valid_ids and r not in valid_ids]
                supported = [r for r in refs if (not valid_ids) or (r in valid_ids)]
                if invalid:
                    mapping_note = ""
                    if mapped_pairs:
                        mapping_note = (
                            " | numeric REC mapping: "
                            + ", ".join(f"{src}->{dst}" for src, dst in mapped_pairs)
                        )
                    findings.append(
                        _brief_finding(
                            run_id,
                            brief_id,
                            section=section,
                            claim_text=text,
                            supported_record_ids=supported,
                            grounded_to_records=False,
                            issue_type="rec_mismatch",
                            severity="High",
                            notes=f"REC reference(s) not in selected_record_ids: {invalid}{mapping_note}",
                        )
                    )
            elif section in _CLAIM_HEADINGS and _is_bullet_line(text):
                if _is_structural_topic_label_bullet(text, section):
                    continue
                findings.append(
                    _brief_finding(
                        run_id,
                        brief_id,
                        section=section,
                        claim_text=text,
                        supported_record_ids=[],
                        grounded_to_records=False,
                        issue_type="ungrounded_claim",
                        severity="High",
                        notes="Bullet claim has no REC citation while brief uses REC labels.",
                    )
                )

    uncertainty_required = any(_record_contains_uncertainty(r) for r in selected_records)
    uncertainty_text = sections.get("CONFLICTS & UNCERTAINTY", "")
    uncertainty_non_empty = bool(
        uncertainty_text
        and not re.search(r"\bnone observed\b|\bnone\b|\bn/?a\b", uncertainty_text, re.IGNORECASE)
    )
    if uncertainty_required and not uncertainty_non_empty:
        findings.append(
            _brief_finding(
                run_id,
                brief_id,
                section="CONFLICTS & UNCERTAINTY",
                claim_text=uncertainty_text or "(empty)",
                supported_record_ids=[],
                grounded_to_records=False,
                issue_type="missing_uncertainty",
                severity="High",
                notes="At least one selected record has uncertainty language; section is empty or 'None observed'.",
            )
        )

    if uncertainty_required and _records_have_soft_language(selected_records):
        for m in _OVERREACH_RE.finditer(brief_text or ""):
            sentence = _extract_sentence(brief_text, m.start())
            if not sentence:
                continue
            findings.append(
                _brief_finding(
                    run_id,
                    brief_id,
                    section="EXECUTIVE SUMMARY",
                    claim_text=sentence,
                    supported_record_ids=[],
                    grounded_to_records=False,
                    issue_type="overreach",
                    severity="Medium",
                    notes="Overreach wording found while source records include softer uncertainty language.",
                )
            )
            if sum(1 for f in findings if f.get("issue_type") == "overreach") >= 3:
                break

    cross_record_count = _cross_record_theme_count(lined, valid_ids)
    action_score = _action_specificity_score(sections.get("RECOMMENDED ACTIONS", ""))

    metrics = {
        "uncertainty_required": uncertainty_required,
        "uncertainty_compliant": (not uncertainty_required) or uncertainty_non_empty,
        "cross_record_theme_count": cross_record_count,
        "action_specificity_score": action_score,
    }
    for finding in findings:
        finding["version"] = run_version
    return findings, metrics


def _severity_counts(rows: Sequence[Dict[str, Any]]) -> Dict[str, int]:
    counts = {"High": 0, "Medium": 0, "Low": 0}
    for row in rows:
        sev = str(row.get("severity") or "").title()
        if sev in counts:
            counts[sev] += 1
    return counts


def _weighted_record_score(counts: Dict[str, int]) -> int:
    score = 100 - 25 * counts.get("High", 0) - 10 * counts.get("Medium", 0) - 2 * counts.get("Low", 0)
    return max(0, int(score))


def _per_record_scores(
    findings: Sequence[Dict[str, Any]],
    all_record_ids: Sequence[str],
) -> Dict[str, Dict[str, Any]]:
    """Compute per-record QC score and finding breakdown.

    Returns {record_id: {"score": int, "high": int, "medium": int, "low": int, "findings": [...]}}
    Records with no findings get score=100.
    """
    by_rid: Dict[str, List[Dict[str, Any]]] = {}
    for rid in all_record_ids:
        by_rid.setdefault(rid, [])
    for f in findings:
        rid = str(f.get("record_id") or "")
        if rid:
            by_rid.setdefault(rid, []).append(f)
    out: Dict[str, Dict[str, Any]] = {}
    for rid, rec_findings in by_rid.items():
        counts = _severity_counts(rec_findings)
        out[rid] = {
            "score": _weighted_record_score(counts),
            "high": counts["High"],
            "medium": counts["Medium"],
            "low": counts["Low"],
            "findings": rec_findings,
        }
    return out


def _avg_record_score(per_record: Dict[str, Dict[str, Any]]) -> float:
    """Average of per-record QC scores."""
    if not per_record:
        return 100.0
    return round(sum(r["score"] for r in per_record.values()) / len(per_record), 1)


def _weighted_brief_score(brief_findings: Sequence[Dict[str, Any]], cross_record_theme_count: int) -> int:
    by_issue = Counter(str(r.get("issue_type") or "") for r in brief_findings)
    ungrounded = by_issue.get("ungrounded_claim", 0) + by_issue.get("rec_mismatch", 0)
    wrong_signal = by_issue.get("wrong_certainty", 0) + by_issue.get("geo_distortion", 0) + by_issue.get("rec_mismatch", 0)
    overreach = by_issue.get("overreach", 0)
    missing_uncertainty = 1 if by_issue.get("missing_uncertainty", 0) > 0 else 0
    missing_cross_record = 1 if cross_record_theme_count < 2 else 0
    score = (
        100
        - 25 * ungrounded
        - 20 * wrong_signal
        - 10 * overreach
        - 20 * missing_uncertainty
        - 15 * missing_cross_record
    )
    return max(0, int(score))


def _col_letter(idx: int) -> str:
    out = ""
    n = max(1, idx)
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def _sheet_xml_from_df(df: pd.DataFrame) -> str:
    columns = [str(c) for c in df.columns]
    rows: List[List[Any]] = [columns]
    if not df.empty:
        rows.extend(df.where(pd.notna(df), "").values.tolist())

    col_count = max(1, len(columns))
    row_count = max(1, len(rows))
    last_cell = f"{_col_letter(col_count)}{row_count}"

    widths: List[float] = []
    for col_idx in range(col_count):
        if col_idx < len(columns):
            max_len = len(columns[col_idx])
            sample_len = min(len(rows), 300)
            for r in range(1, sample_len):
                if col_idx >= len(rows[r]):
                    continue
                max_len = max(max_len, len(str(rows[r][col_idx] if rows[r][col_idx] is not None else "")))
        else:
            max_len = 12
        widths.append(float(max(12, min(60, max_len + 2))))

    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        f'<dimension ref="A1:{last_cell}"/>',
        '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>',
        "<cols>",
    ]
    for i, w in enumerate(widths, start=1):
        parts.append(f'<col min="{i}" max="{i}" width="{w:.2f}" customWidth="1"/>')
    parts.append("</cols>")
    parts.append("<sheetData>")

    for r_idx, row in enumerate(rows, start=1):
        cells: List[str] = []
        for c_idx in range(col_count):
            value = row[c_idx] if c_idx < len(row) else ""
            if value is None or value == "":
                continue
            ref = f"{_col_letter(c_idx + 1)}{r_idx}"
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cells.append(f'<c r="{ref}" s="1"><v>{value}</v></c>')
            else:
                txt = xml_escape(str(value))
                cells.append(f'<c r="{ref}" s="1" t="inlineStr"><is><t>{txt}</t></is></c>')
        if cells:
            parts.append(f'<row r="{r_idx}">{"".join(cells)}</row>')
    parts.append("</sheetData></worksheet>")
    return "".join(parts)


def _write_minimal_xlsx(report_path: Path, sheets: Sequence[Tuple[str, pd.DataFrame]]) -> None:
    safe_sheets: List[Tuple[str, pd.DataFrame]] = []
    used: set[str] = set()
    for idx, (name, df) in enumerate(sheets, start=1):
        base = (name or f"Sheet{idx}")[:31]
        cand = base
        suffix = 1
        while cand in used:
            suffix += 1
            cand = f"{base[:28]}_{suffix}"
        used.add(cand)
        safe_sheets.append((cand, df))

    workbook_sheets = []
    workbook_rels = []
    content_overrides = []
    worksheet_xml: Dict[str, str] = {}

    for i, (name, df) in enumerate(safe_sheets, start=1):
        sheet_file = f"sheet{i}.xml"
        rid = f"rId{i}"
        workbook_sheets.append(f'<sheet name="{xml_escape(name)}" sheetId="{i}" r:id="{rid}"/>')
        workbook_rels.append(
            f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/{sheet_file}"/>'
        )
        content_overrides.append(
            f'<Override PartName="/xl/worksheets/{sheet_file}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
        worksheet_xml[sheet_file] = _sheet_xml_from_df(df)

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets>{"".join(workbook_sheets)}</sheets>'
        "</workbook>"
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(workbook_rels)}'
        '<Relationship Id="rIdStyles" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        "</Relationships>"
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="2">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment wrapText="1" vertical="top"/></xf>'
        "</cellXfs>"
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        f'{"".join(content_overrides)}'
        "</Types>"
    )
    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        "</Relationships>"
    )
    now_iso = _now().strftime("%Y-%m-%dT%H:%M:%SZ")
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<dc:creator>Cognitra</dc:creator>"
        "<cp:lastModifiedBy>Cognitra</cp:lastModifiedBy>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{now_iso}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{now_iso}</dcterms:modified>'
        "</cp:coreProperties>"
    )
    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        "<Application>Cognitra</Application>"
        "</Properties>"
    )

    with zipfile.ZipFile(report_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", root_rels_xml)
        zf.writestr("docProps/core.xml", core_xml)
        zf.writestr("docProps/app.xml", app_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        zf.writestr("xl/styles.xml", styles_xml)
        for fname, xml in worksheet_xml.items():
            zf.writestr(f"xl/worksheets/{fname}", xml)


def export_quality_excel(report_path: Path = QUALITY_REPORT_XLSX) -> Path:
    _ensure_quality_dir()
    record_rows = _read_jsonl(RECORD_QC_LOG)
    brief_rows = _read_jsonl(BRIEF_QC_LOG)
    run_rows = _read_jsonl(QUALITY_RUNS_LOG)

    record_df = pd.DataFrame(record_rows)
    brief_df = pd.DataFrame(brief_rows)
    runs_df = pd.DataFrame(run_rows)

    summary_rows: List[Dict[str, Any]] = []
    if not record_df.empty:
        if "severity" in record_df.columns:
            for k, v in record_df["severity"].value_counts().to_dict().items():
                summary_rows.append({"category": "record_by_severity", "key": k, "count": int(v)})
        if "finding_type" in record_df.columns:
            for k, v in record_df["finding_type"].value_counts().to_dict().items():
                summary_rows.append({"category": "record_by_finding_type", "key": k, "count": int(v)})
        if "record_id" in record_df.columns:
            for k, v in record_df["record_id"].value_counts().head(20).to_dict().items():
                summary_rows.append({"category": "record_by_record_id", "key": k, "count": int(v)})

    if not brief_df.empty:
        if "severity" in brief_df.columns:
            for k, v in brief_df["severity"].value_counts().to_dict().items():
                summary_rows.append({"category": "brief_by_severity", "key": k, "count": int(v)})
        if "issue_type" in brief_df.columns:
            for k, v in brief_df["issue_type"].value_counts().to_dict().items():
                summary_rows.append({"category": "brief_by_issue_type", "key": k, "count": int(v)})
        if "brief_id" in brief_df.columns:
            for k, v in brief_df["brief_id"].value_counts().head(20).to_dict().items():
                summary_rows.append({"category": "brief_by_brief_id", "key": k, "count": int(v)})

    pivot_df = pd.DataFrame(summary_rows)

    # Build KPI trends sheet from runs data
    trends_rows: List[Dict[str, Any]] = []
    if not runs_df.empty and len(runs_df) >= 2:
        for kpi in _KPI_KEYS:
            if kpi not in runs_df.columns:
                continue
            vals = runs_df[kpi].dropna().tolist()
            if len(vals) < 2:
                continue
            current = float(vals[-1])
            previous = float(vals[-2])
            avg_all = sum(float(v) for v in vals) / len(vals)
            delta = current - previous
            direction_pref = _KPI_DIRECTION.get(kpi, "higher_is_better")
            if direction_pref == "higher_is_better":
                trend_dir = "improving" if delta > 0 else ("declining" if delta < 0 else "stable")
            else:
                trend_dir = "improving" if delta < 0 else ("declining" if delta > 0 else "stable")
            trends_rows.append({
                "kpi": kpi,
                "current": round(current, 4),
                "previous": round(previous, 4),
                "delta": round(delta, 4),
                "avg_all_runs": round(avg_all, 4),
                "direction": trend_dir,
                "n_runs": len(vals),
            })
    trends_df = pd.DataFrame(trends_rows)

    sheets = [
        ("record_qc", record_df),
        ("brief_qc", brief_df),
        ("runs_summary", runs_df),
        ("kpi_trends", trends_df),
        ("summary_pivot", pivot_df),
    ]

    if importlib.util.find_spec("openpyxl") or importlib.util.find_spec("xlsxwriter"):
        engine = "openpyxl" if importlib.util.find_spec("openpyxl") else "xlsxwriter"
        with pd.ExcelWriter(report_path, engine=engine) as writer:
            for sheet_name, df in sheets:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            if engine == "openpyxl":
                try:
                    from openpyxl.styles import Alignment
                    from openpyxl.utils import get_column_letter

                    for ws in writer.book.worksheets:
                        ws.freeze_panes = "A2"
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                        for col_idx in range(1, ws.max_column + 1):
                            col_letter = get_column_letter(col_idx)
                            max_len = 0
                            for row_idx in range(1, min(ws.max_row, 300) + 1):
                                val = ws.cell(row=row_idx, column=col_idx).value
                                if val is None:
                                    continue
                                max_len = max(max_len, len(str(val)))
                            ws.column_dimensions[col_letter].width = max(12, min(60, max_len + 2))
                except Exception:
                    pass
    else:
        _write_minimal_xlsx(report_path, sheets)

    return report_path


_KPI_KEYS = [
    "KPI-R1", "KPI-R2", "KPI-R3", "KPI-R4", "KPI-R5",
    "KPI-B1", "KPI-B2", "KPI-B3", "KPI-B4", "KPI-B5",
    "weighted_record_score", "weighted_brief_score", "weighted_overall_score",
]

# KPI direction: "lower_is_better" for error rates/counts, "higher_is_better" for pass rates/scores
_KPI_DIRECTION: Dict[str, str] = {
    "KPI-R1": "lower_is_better",   # high severity rate
    "KPI-R2": "lower_is_better",   # medium severity rate
    "KPI-R3": "higher_is_better",  # evidence grounding pass rate
    "KPI-R4": "higher_is_better",  # canonicalization pass rate
    "KPI-R5": "higher_is_better",  # geo pass rate
    "KPI-B1": "lower_is_better",   # ungrounded claims count
    "KPI-B2": "lower_is_better",   # overreach count
    "KPI-B3": "higher_is_better",  # uncertainty compliance
    "KPI-B4": "higher_is_better",  # cross-record theme count
    "KPI-B5": "higher_is_better",  # action specificity score
    "weighted_record_score": "higher_is_better",
    "weighted_brief_score": "higher_is_better",
    "weighted_overall_score": "higher_is_better",
}

_REGRESSION_THRESHOLD = 0.05  # 5% change triggers alert for ratio KPIs
_REGRESSION_ABS_THRESHOLD = 2  # absolute change for count KPIs


def compute_quality_trends(
    current_run: Dict[str, Any],
    lookback: int = 5,
) -> Dict[str, Any]:
    """Compute KPI trends by comparing current run against previous runs."""
    prior_runs = _read_jsonl(QUALITY_RUNS_LOG)
    # Exclude the current run if already appended
    prior_runs = [r for r in prior_runs if r.get("run_id") != current_run.get("run_id")]
    prior_runs = prior_runs[-lookback:] if len(prior_runs) > lookback else prior_runs

    if not prior_runs:
        return {"trends": {}, "regression_alerts": [], "prior_run_count": 0}

    trends: Dict[str, Dict[str, Any]] = {}
    regression_alerts: List[str] = []

    for kpi in _KPI_KEYS:
        current_val = current_run.get(kpi)
        if current_val is None:
            continue
        current_val = float(current_val)

        prior_vals = []
        for r in prior_runs:
            v = r.get(kpi)
            if v is not None:
                prior_vals.append(float(v))
        if not prior_vals:
            continue

        avg_prior = sum(prior_vals) / len(prior_vals)
        last_val = prior_vals[-1]
        delta_from_avg = current_val - avg_prior
        delta_from_last = current_val - last_val

        direction_pref = _KPI_DIRECTION.get(kpi, "higher_is_better")
        if direction_pref == "higher_is_better":
            trend_dir = "improving" if delta_from_avg > 0 else ("declining" if delta_from_avg < 0 else "stable")
            is_regression = delta_from_avg < 0
        else:
            trend_dir = "improving" if delta_from_avg < 0 else ("declining" if delta_from_avg > 0 else "stable")
            is_regression = delta_from_avg > 0

        trends[kpi] = {
            "current": current_val,
            "last": last_val,
            "avg_prior": round(avg_prior, 4),
            "delta_from_avg": round(delta_from_avg, 4),
            "delta_from_last": round(delta_from_last, 4),
            "direction": trend_dir,
        }

        # Check regression threshold
        if is_regression:
            is_ratio = kpi in {"KPI-R1", "KPI-R2", "KPI-R3", "KPI-R4", "KPI-R5", "KPI-B3"}
            if is_ratio and abs(delta_from_avg) >= _REGRESSION_THRESHOLD:
                regression_alerts.append(
                    f"{kpi} {trend_dir}: {avg_prior:.0%} -> {current_val:.0%} (delta {delta_from_avg:+.0%})"
                )
            elif not is_ratio and abs(delta_from_avg) >= _REGRESSION_ABS_THRESHOLD:
                regression_alerts.append(
                    f"{kpi} {trend_dir}: {avg_prior:.1f} -> {current_val:.1f} (delta {delta_from_avg:+.1f})"
                )

    return {
        "trends": trends,
        "regression_alerts": regression_alerts,
        "prior_run_count": len(prior_runs),
    }


# ---------------------------------------------------------------------------
# Feedback loop: aggregate chronic issues into actionable suggestions
# ---------------------------------------------------------------------------

_FINDING_TYPE_SUGGESTIONS: Dict[str, str] = {
    "evidence_not_grounded": (
        "Evidence bullets frequently fail grounding check. "
        "Consider adding 'Use exact phrases from the source text where possible' to the extraction prompt."
    ),
    "evidence_near_miss": (
        "Evidence bullets are often paraphrased (near-miss grounding). "
        "Consider instructing the LLM to quote key facts verbatim rather than paraphrasing."
    ),
    "missing_footprint_region": (
        "Country mentions are present but derived footprint regions are missing. "
        "The extraction prompt may need explicit examples of country-to-region mapping."
    ),
    "display_bucket_leakage": (
        "Country-level values are leaking into the display regions field. "
        "Reinforce in the extraction prompt that regions_mentioned should use broad geographic buckets only."
    ),
    "canonicalization_inconsistent": (
        "Multiple name variants for the same company appear within single records. "
        "Consider expanding the company canonicalization map in postprocess.py."
    ),
    "canonicalization_alias_present": (
        "Non-canonical company name forms are common. "
        "Extend _OEM_CANONICAL_BY_LOWER or _COMPANY_SPECIAL_CANONICAL in postprocess.py."
    ),
    "missing_evidence": (
        "Some records have no evidence bullets at all. "
        "Ensure the extraction prompt enforces minItems=2 for evidence_bullets."
    ),
    "confidence_evidence_mismatch": (
        "High confidence is assigned despite thin evidence. "
        "Review _compute_confidence() signal weights — evidence_bullets threshold may be too lenient."
    ),
    "missing_priority_reason": (
        "Priority escalations lack explanations. "
        "Ensure _boost_priority() and _escalate_priority_from_macro_themes() always set priority_reason."
    ),
    "duplicate_record_in_brief": (
        "Identical stories are being ingested as separate records. "
        "Consider tightening the dedup key or lowering the similarity threshold at ingest time."
    ),
    "near_duplicate_titles": (
        "Near-duplicate titles are slipping through ingest dedup. "
        "Consider lowering find_similar_title_records() threshold from 0.88 to 0.85."
    ),
    "macro_theme_min_groups_violation": (
        "Macro themes are firing without meeting minimum signal group thresholds. "
        "Investigate _detect_macro_themes() logic for edge cases."
    ),
}


def generate_extraction_feedback(
    lookback_runs: int = 5,
) -> Dict[str, Any]:
    """Aggregate findings across recent runs to identify chronic issues and suggest prompt improvements."""
    run_rows = _read_jsonl(QUALITY_RUNS_LOG)
    recent_run_ids = {r.get("run_id") for r in run_rows[-lookback_runs:]} if run_rows else set()

    if not recent_run_ids:
        return {"chronic_issues": [], "prompt_suggestions": [], "runs_analyzed": 0}

    record_rows = _read_jsonl(RECORD_QC_LOG)
    brief_rows = _read_jsonl(BRIEF_QC_LOG)

    # Count finding types across recent runs
    type_by_run: Dict[str, set] = {}  # finding_type -> set of run_ids it appeared in
    type_examples: Dict[str, List[str]] = {}  # finding_type -> sample notes

    for row in record_rows:
        rid = row.get("run_id")
        if rid not in recent_run_ids:
            continue
        ft = str(row.get("finding_type") or "")
        if not ft:
            continue
        type_by_run.setdefault(ft, set()).add(rid)
        examples = type_examples.setdefault(ft, [])
        if len(examples) < 3:
            notes = str(row.get("notes") or "")[:120]
            if notes:
                examples.append(notes)

    for row in brief_rows:
        rid = row.get("run_id")
        if rid not in recent_run_ids:
            continue
        it = str(row.get("issue_type") or "")
        if not it:
            continue
        type_by_run.setdefault(it, set()).add(rid)
        examples = type_examples.setdefault(it, [])
        if len(examples) < 3:
            notes = str(row.get("notes") or "")[:120]
            if notes:
                examples.append(notes)

    n_runs = len(recent_run_ids)
    chronic_issues: List[Dict[str, Any]] = []
    prompt_suggestions: List[str] = []

    # Chronic = appears in >= 50% of recent runs
    for ft, run_ids in sorted(type_by_run.items(), key=lambda x: len(x[1]), reverse=True):
        frequency = len(run_ids)
        if frequency < max(1, n_runs // 2):
            continue
        chronic_issues.append({
            "type": ft,
            "frequency": f"{frequency} of {n_runs} runs",
            "frequency_ratio": round(frequency / n_runs, 2),
            "example_notes": type_examples.get(ft, []),
        })
        suggestion = _FINDING_TYPE_SUGGESTIONS.get(ft)
        if suggestion and suggestion not in prompt_suggestions:
            prompt_suggestions.append(suggestion)

    return {
        "chronic_issues": chronic_issues,
        "prompt_suggestions": prompt_suggestions,
        "runs_analyzed": n_runs,
    }


def run_record_only_qc(
    *,
    record_ids: Optional[Sequence[str]] = None,
) -> Dict[str, Any]:
    """Run only record-level QC checks (no brief QC)."""
    _ensure_quality_dir()
    run_version = _next_run_version()
    run_id = _new_run_id()
    created_at = _now_iso()

    all_records = load_records()
    target_mode = "all_records"
    target_records = all_records
    if record_ids:
        wanted = {str(x) for x in record_ids if str(x)}
        target_mode = "explicit_record_ids"
        target_records = [
            r for r in all_records if str(r.get("record_id") or "") in wanted
        ]

    record_findings, record_metrics = run_record_qc(run_id, target_records, run_version)
    _append_jsonl(RECORD_QC_LOG, record_findings)

    rec_counts = _severity_counts(record_findings)
    n_records = len(target_records)

    evidence_pass = record_metrics.get("evidence_pass") or {}
    canonical_pass = record_metrics.get("canonical_pass") or {}
    geo_pass = record_metrics.get("geo_pass") or {}

    kpi_r1 = (rec_counts["High"] / n_records) if n_records else 0.0
    kpi_r2 = (rec_counts["Medium"] / n_records) if n_records else 0.0
    kpi_r3 = (sum(1 for v in evidence_pass.values() if v) / n_records) if n_records else 0.0
    kpi_r4 = (sum(1 for v in canonical_pass.values() if v) / n_records) if n_records else 0.0
    kpi_r5 = (sum(1 for v in geo_pass.values() if v) / n_records) if n_records else 0.0

    per_record = record_metrics.get("per_record_scores") or {}
    weighted_record = round(_avg_record_score(per_record))

    run_summary = {
        "run_id": run_id,
        "run_version": run_version,
        "created_at": created_at,
        "qc_scope": "record_only",
        "brief_id": None,
        "week_range": None,
        "target_mode": target_mode,
        "target_record_count": n_records,
        "record_qc_high": rec_counts["High"],
        "record_qc_medium": rec_counts["Medium"],
        "record_qc_low": rec_counts["Low"],
        "brief_qc_high": None,
        "brief_qc_medium": None,
        "brief_qc_low": None,
        "KPI-R1": round(kpi_r1, 4),
        "KPI-R2": round(kpi_r2, 4),
        "KPI-R3": round(kpi_r3, 4),
        "KPI-R4": round(kpi_r4, 4),
        "KPI-R5": round(kpi_r5, 4),
        "KPI-B1": None,
        "KPI-B2": None,
        "KPI-B3": None,
        "KPI-B4": None,
        "KPI-B5": None,
        "weighted_record_score": weighted_record,
        "weighted_brief_score": None,
        "weighted_overall_score": weighted_record,
    }
    _append_jsonl(QUALITY_RUNS_LOG, [run_summary])

    trend_result = compute_quality_trends(run_summary)
    feedback = generate_extraction_feedback(lookback_runs=5)
    report_path = export_quality_excel()

    issue_counter = Counter()
    for row in record_findings:
        issue_counter[str(row.get("finding_type") or "")] += 1
    top_issue_types = [k for k, _ in issue_counter.most_common(3)]

    return {
        "run_id": run_id,
        "run_version": run_version,
        "qc_scope": "record_only",
        "brief_id": None,
        "week_range": None,
        "target_record_count": n_records,
        "record_counts": rec_counts,
        "brief_counts": {"High": 0, "Medium": 0, "Low": 0},
        "weighted_record_score": weighted_record,
        "weighted_brief_score": None,
        "per_record_scores": per_record,
        "top_issue_types": top_issue_types,
        "report_path": str(report_path),
        "run_summary": run_summary,
        "trends": trend_result,
        "feedback": feedback,
    }


def run_quality_pipeline(
    *,
    brief_id: Optional[str] = None,
    use_latest_brief: bool = True,
) -> Dict[str, Any]:
    _ensure_quality_dir()
    run_version = _next_run_version()
    run_id = _new_run_id()
    created_at = _now_iso()

    records = load_records()
    brief_path = _brief_file_from_id(brief_id) if brief_id else (_latest_brief_file() if use_latest_brief else None)
    brief_meta = _brief_meta_for_file(brief_path)
    brief_text = brief_path.read_text(encoding="utf-8") if brief_path and brief_path.exists() else ""
    resolved_brief_id = brief_path.stem if brief_path else ""

    selected_ids = [str(x) for x in (brief_meta.get("selected_record_ids") or []) if str(x)]
    week_range = str(brief_meta.get("week_range") or "Last 7 days")
    week_days = _parse_week_range_days(week_range, default_days=7)

    target_records, target_mode = _select_target_records(records, selected_ids, week_days)
    selected_records = target_records

    record_findings, record_metrics = run_record_qc(run_id, target_records, run_version)
    brief_findings, brief_metrics = run_brief_qc(
        run_id=run_id,
        run_version=run_version,
        brief_id=resolved_brief_id or "no_brief",
        brief_text=brief_text,
        selected_record_ids=selected_ids,
        selected_records=selected_records,
    )

    _append_jsonl(RECORD_QC_LOG, record_findings)
    _append_jsonl(BRIEF_QC_LOG, brief_findings)

    rec_counts = _severity_counts(record_findings)
    brief_counts = _severity_counts(brief_findings)
    n_records = len(target_records)

    evidence_pass = record_metrics.get("evidence_pass") or {}
    canonical_pass = record_metrics.get("canonical_pass") or {}
    geo_pass = record_metrics.get("geo_pass") or {}

    kpi_r1 = (rec_counts["High"] / n_records) if n_records else 0.0
    kpi_r2 = (rec_counts["Medium"] / n_records) if n_records else 0.0
    kpi_r3 = (sum(1 for v in evidence_pass.values() if v) / n_records) if n_records else 0.0
    kpi_r4 = (sum(1 for v in canonical_pass.values() if v) / n_records) if n_records else 0.0
    kpi_r5 = (sum(1 for v in geo_pass.values() if v) / n_records) if n_records else 0.0

    by_issue = Counter(str(r.get("issue_type") or "") for r in brief_findings)
    kpi_b1 = by_issue.get("ungrounded_claim", 0) + by_issue.get("rec_mismatch", 0)
    kpi_b2 = by_issue.get("overreach", 0)
    kpi_b3 = 1.0 if brief_metrics.get("uncertainty_compliant", True) else 0.0
    kpi_b4 = int(brief_metrics.get("cross_record_theme_count") or 0)
    kpi_b5 = int(brief_metrics.get("action_specificity_score") or 1)

    per_record = record_metrics.get("per_record_scores") or {}
    weighted_record = round(_avg_record_score(per_record))
    weighted_brief = _weighted_brief_score(brief_findings, cross_record_theme_count=kpi_b4)

    run_summary = {
        "run_id": run_id,
        "run_version": run_version,
        "created_at": created_at,
        "brief_id": resolved_brief_id,
        "week_range": week_range,
        "target_mode": target_mode,
        "target_record_count": n_records,
        "record_qc_high": rec_counts["High"],
        "record_qc_medium": rec_counts["Medium"],
        "record_qc_low": rec_counts["Low"],
        "brief_qc_high": brief_counts["High"],
        "brief_qc_medium": brief_counts["Medium"],
        "brief_qc_low": brief_counts["Low"],
        "KPI-R1": round(kpi_r1, 4),
        "KPI-R2": round(kpi_r2, 4),
        "KPI-R3": round(kpi_r3, 4),
        "KPI-R4": round(kpi_r4, 4),
        "KPI-R5": round(kpi_r5, 4),
        "KPI-B1": int(kpi_b1),
        "KPI-B2": int(kpi_b2),
        "KPI-B3": round(kpi_b3, 4),
        "KPI-B4": int(kpi_b4),
        "KPI-B5": int(kpi_b5),
        "weighted_record_score": weighted_record,
        "weighted_brief_score": weighted_brief,
        "weighted_overall_score": round((weighted_record + weighted_brief) / 2.0, 1),
    }
    _append_jsonl(QUALITY_RUNS_LOG, [run_summary])

    # --- Trend analysis ---
    trend_result = compute_quality_trends(run_summary)

    # --- Feedback loop ---
    feedback = generate_extraction_feedback(lookback_runs=5)

    report_path = export_quality_excel()

    issue_counter = Counter()
    for row in record_findings:
        issue_counter[str(row.get("finding_type") or "")] += 1
    for row in brief_findings:
        issue_counter[str(row.get("issue_type") or "")] += 1
    top_issue_types = [k for k, _ in issue_counter.most_common(3)]

    return {
        "run_id": run_id,
        "run_version": run_version,
        "brief_id": resolved_brief_id,
        "week_range": week_range,
        "target_record_count": n_records,
        "record_counts": rec_counts,
        "brief_counts": brief_counts,
        "weighted_record_score": weighted_record,
        "weighted_brief_score": weighted_brief,
        "per_record_scores": per_record,
        "top_issue_types": top_issue_types,
        "report_path": str(report_path),
        "run_summary": run_summary,
        "trends": trend_result,
        "feedback": feedback,
    }
